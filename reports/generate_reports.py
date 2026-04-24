# reports/generate_reports.py
import pandas as pd
import psycopg2
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
import warnings

# Suppress the pandas SQLAlchemy warning for a cleaner terminal output
warnings.filterwarnings('ignore', category=UserWarning)

# UPDATE YOUR PASSWORD HERE
CONN_STR = 'postgresql://postgres:12345@localhost:5432/medicare_claims'

def fetch(sql, step_name):
    print(f"Running query: {step_name} (This may take 30-90 seconds...)")
    conn = psycopg2.connect(CONN_STR)
    df = pd.read_sql(sql, conn)
    conn.close()
    return df

def style_header(ws, hex_color='1F4E79'):
    fill = PatternFill('solid', fgColor=hex_color)
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')

def write_df_to_sheet(ws, df):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

print("Starting automated Excel report generation...")

# ── Report 1: State spend summary ──────────────────────────────────────────
df_state = fetch("""
    SELECT state,
           ROUND(SUM(total_drug_cost)/1e6, 2)  AS spend_millions,
           SUM(total_claims)                   AS total_claims,
           COUNT(DISTINCT npi)                 AS provider_count
    FROM claims.part_d_prescribers
    GROUP BY state 
    ORDER BY spend_millions DESC 
    LIMIT 20
""", "State Spend Summary")

# ── Report 2: Top 20 drugs by national spend ───────────────────────────────
df_drugs = fetch("""
    SELECT generic_name,
           ROUND(SUM(total_drug_cost)/1e6, 2)  AS spend_millions,
           SUM(total_claims)                   AS total_claims,
           COUNT(npi)                          AS prescriber_count,
           ROUND(SUM(total_drug_cost)/NULLIF(SUM(total_claims),0), 2) AS avg_cost_per_claim
    FROM claims.part_d_prescribers
    GROUP BY generic_name 
    ORDER BY spend_millions DESC 
    LIMIT 20
""", "Top Drugs by National Spend")

# ── Report 3: Aberrant Providers in SC (Tailored to JD) ─────────────────────
df_outliers = fetch("""
    WITH stats AS (
        SELECT npi, provider_name, state, specialty,
               SUM(total_claims) AS claims,
               ROUND(SUM(total_drug_cost)/NULLIF(SUM(total_claims),0),2) AS cpc
        FROM claims.part_d_prescribers
        WHERE state = 'SC'
        GROUP BY npi, provider_name, state, specialty
        HAVING SUM(total_claims) > 100
    ),
    dist AS (SELECT AVG(cpc) AS m, STDDEV(cpc) AS s_dev FROM stats)
    SELECT stats.*, 
           ROUND((stats.cpc - dist.m)/NULLIF(dist.s_dev,0), 3) AS z_score,
           CASE WHEN ABS((stats.cpc - dist.m)/NULLIF(dist.s_dev,0)) > 3 THEN 'HIGH OUTLIER'
                WHEN ABS((stats.cpc - dist.m)/NULLIF(dist.s_dev,0)) > 2 THEN 'MODERATE OUTLIER' 
           END AS flag
    FROM stats, dist
    WHERE ABS((stats.cpc - dist.m)/NULLIF(dist.s_dev,0)) > 2
    ORDER BY ABS((stats.cpc - dist.m)/NULLIF(dist.s_dev,0)) DESC 
    LIMIT 100
""", "Aberrant Providers in SC")

# ── Build workbook ─────────────────────────────────────────────────────────
print("Building Excel Workbook & Charts...")
wb = Workbook()
for title, df in [('State Spend Summary', df_state),
                  ('Top Drugs National',  df_drugs),
                  ('Aberrant Providers SC',  df_outliers)]:
    ws = wb.create_sheet(title)
    write_df_to_sheet(ws, df)

# ── Embed chart in State sheet ─────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(10, 5))
# Plotting the top 15 states in reverse order so largest is on top
ax.barh(df_state['state'][:15][::-1],
        df_state['spend_millions'][:15][::-1],
        color='#1F4E79')
ax.set_xlabel('Spend ($M)')
ax.set_title('Top 15 States — Medicare Part D Drug Spend')
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}M'))
plt.tight_layout()

img_buf = io.BytesIO()
fig.savefig(img_buf, format='png', dpi=120)
img_buf.seek(0)
plt.close()

from openpyxl.drawing.image import Image as XLImage
img = XLImage(img_buf)
img.anchor = 'H2'
wb['State Spend Summary'].add_image(img)

del wb['Sheet']  # remove default blank sheet

# ── Save Output ────────────────────────────────────────────────────────────
# Make sure we save it in the correct path relative to where we run the script
output_path = os.path.join(os.path.dirname(__file__), 'medicare_claims_reports.xlsx')
wb.save(output_path)
print(f"Success! Reports beautifully formatted and written to: {output_path}")